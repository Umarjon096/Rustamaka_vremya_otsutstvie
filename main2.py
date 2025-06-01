import pandas as pd
from datetime import datetime, date, time, timedelta
from openpyxl.styles import Border, Side

def working_time(file_path, output_path, same_file,
                 work_start_hours, work_end_hours,
                 lunch_start_hours, lunch_end_hours,
                 enter_ips, exit_ips):
    # 1. Считываем исходный Excel-файл
    df = pd.read_excel(file_path)

    # 2. Переименуем колонки (если нужно)
    df = df.rename(columns={
        'Точка СКУД': 'СКУД',
        'Дата': 'Дата',
        'Время': 'Время',
        'Name': 'Сотрудник'
    })

    # 3. Собираем единый datetime-столбец
    df['Дата и время'] = pd.to_datetime(
        df['Дата'].astype(str) + ' ' + df['Время'].astype(str),
        dayfirst=True,
        errors='coerce'
    )

    # 4. Обрезаем пробелы и префиксы в URL-ах, а затем определяем «Действие» – «Вход»/«Выход»/«Неизвестно»
    df['СКУД'] = (
        df['СКУД']
        .astype(str)
        .str.strip()
        .str.replace(r'^https?://', '', regex=True)  # убираем http:// или https://, если есть
    )

    entry_points = enter_ips.strip().split(',')  # например: "10.100.6.65,10.100.6.79"
    exit_points  = exit_ips.strip().split(',')   # например: "10.100.6.25,10.100.6.38"

    df['Действие'] = df['СКУД'].apply(
        lambda url: 'Вход' if any(ip in url for ip in entry_points)
        else ('Выход' if any(ip in url for ip in exit_points) else 'Неизвестно')
    )

    # 5. Разбиваем обратно на отдельные столбцы «Дата» и «Время» (для группировки)
    df['Дата'] = df['Дата и время'].dt.date
    df['Время'] = df['Дата и время'].dt.time

    # 6. Сортируем по Сотруднику и времени
    df = df.sort_values(by=['Сотрудник', 'Дата и время']).reset_index(drop=True)

    # =========================
    # 7. Фильтрация подряд идущих одинаковых событий
    # =========================
    # Оставляем только первое событие из каждой серии «Вход»-«Вход»-… или «Выход»-«Выход»-…
    df['PrevAction'] = df.groupby(['Сотрудник', 'Дата'])['Действие'].shift()
    df = df[
        (df['PrevAction'].isna()) |
        (df['Действие'] != df['PrevAction'])
    ].copy()
    df.drop(columns=['PrevAction'], inplace=True)

    # =========================
    # 7a. Считаем количество «Вход» и «Выход» (после фильтрации подрядных дубликатов)
    # =========================
    counts = (
        df[df['Действие'].isin(['Вход', 'Выход'])]
        .groupby(['Сотрудник', 'Дата'])['Действие']
        .value_counts()
        .unstack(fill_value=0)
        .reset_index()
    )
    # Переименуем колонки для удобства:
    counts = counts.rename(columns={
        'Вход': 'Кол-во входов',
        'Выход': 'Кол-во выходов'
    })
    # Если какого-то действия не было, ещё не будет в counts, добавим нули
    if 'Кол-во входов' not in counts:
        counts['Кол-во входов'] = 0
    if 'Кол-во выходов' not in counts:
        counts['Кол-во выходов'] = 0

    # =========================
    # 8. Находим время прихода и ухода
    # =========================
    enter = (
        df[df['Действие'] == 'Вход']
        .groupby(['Сотрудник', 'Дата'], as_index=False)['Дата и время']
        .min()
        .rename(columns={'Дата и время': 'Время прихода'})
    )
    enter['Время прихода'] = enter['Время прихода'].dt.time

    exit_ = (
        df[df['Действие'] == 'Выход']
        .groupby(['Сотрудник', 'Дата'], as_index=False)['Дата и время']
        .max()
        .rename(columns={'Дата и время': 'Время ухода'})
    )
    exit_['Время ухода'] = exit_['Время ухода'].dt.time

    summary = pd.merge(
        enter, exit_, on=['Сотрудник', 'Дата'], how='outer'
    ).sort_values(by=['Сотрудник', 'Дата']).reset_index(drop=True)

    # =========================
    # 9. Считаем «часы отсутствия» (с учётом обеда)
    # =========================
    def adjust_for_lunch(start_abs: datetime, end_abs: datetime) -> timedelta:
        if end_abs <= start_abs:
            return timedelta(0)

        lunch_start = datetime.combine(start_abs.date(), time(lunch_start_hours, 0))
        lunch_end   = datetime.combine(start_abs.date(), time(lunch_end_hours, 0))
        full_interval = end_abs - start_abs

        overlap_start = max(start_abs, lunch_start)
        overlap_end   = min(end_abs, lunch_end)
        lunch_overlap = timedelta(0)
        if overlap_end > overlap_start:
            lunch_overlap = overlap_end - overlap_start

        return full_interval - lunch_overlap

    def compute_absence_within_workday(start_abs: datetime, end_abs: datetime, work_date: date) -> timedelta:
        if end_abs <= start_abs:
            return timedelta(0)

        work_start = datetime.combine(work_date, time(work_start_hours, 0))
        work_end   = datetime.combine(work_date, time(work_end_hours, 0))

        interval_start = max(start_abs, work_start)
        interval_end   = min(end_abs, work_end)

        if interval_end <= interval_start:
            return timedelta(0)

        return adjust_for_lunch(interval_start, interval_end)

    absences = []
    for (employee, work_date), group in df.groupby(['Сотрудник', 'Дата']):
        events = group[group['Действие'].isin(['Вход', 'Выход'])].sort_values('Дата и время')
        total_absence = timedelta(0)

        for idx, row in events.iterrows():
            if row['Действие'] == 'Выход':
                mask_next = (
                    (events['Дата и время'] > row['Дата и время']) &
                    (events['Действие'] == 'Вход')
                )
                if mask_next.any():
                    next_entry = events[mask_next].iloc[0]
                    start_abs = row['Дата и время']
                    end_abs   = next_entry['Дата и время']
                    adjusted = compute_absence_within_workday(start_abs, end_abs, work_date)
                    total_absence += adjusted

        absences.append({
            'Сотрудник': employee,
            'Дата': work_date,
            'Время отсутствия': total_absence
        })

    df_absence = pd.DataFrame(absences)

    # =========================
    # 10. Сводим итоговую таблицу
    # =========================
    result = pd.merge(
        summary, df_absence,
        on=['Сотрудник', 'Дата'],
        how='left'
    ).sort_values(by=['Сотрудник', 'Дата']).reset_index(drop=True)

    # Преобразуем timedelta в строку "ЧЧ:ММ"
    def format_timedelta(td: timedelta) -> str:
        total_seconds = int(td.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"

    result['Часы отсутствия'] = result['Время отсутствия'].apply(
        lambda x: format_timedelta(x) if pd.notnull(x) else "00:00"
    )

    # =========================
    # 11. Добавляем столбец «Переработка»
    # =========================
    work_start_time = time(work_start_hours, 0)
    work_end_time   = time(work_end_hours, 0)

    def compute_overtime(arrival: time, departure: time) -> timedelta:
        overtime = timedelta(0)
        if pd.isna(arrival) or pd.isna(departure):
            return overtime

        if arrival < work_start_time:
            dt_arrival    = datetime.combine(date.today(), arrival)
            dt_work_start = datetime.combine(date.today(), work_start_time)
            overtime += (dt_work_start - dt_arrival)

        if departure > work_end_time:
            dt_departure = datetime.combine(date.today(), departure)
            dt_work_end  = datetime.combine(date.today(), work_end_time)
            overtime += (dt_departure - dt_work_end)

        return overtime

    overtimes = []
    for idx, row in result.iterrows():
        arr = row.get('Время прихода')
        dep = row.get('Время ухода')
        if not isinstance(arr, time):
            arr = pd.NaT
        if not isinstance(dep, time):
            dep = pd.NaT

        ot_delta = compute_overtime(arr, dep)
        overtimes.append(format_timedelta(ot_delta))

    result['Переработка'] = overtimes

    # =========================
    # 12. Объединяем с количеством входов/выходов
    # =========================
    result = pd.merge(
        result,
        counts[['Сотрудник', 'Дата', 'Кол-во входов', 'Кол-во выходов']],
        on=['Сотрудник', 'Дата'],
        how='left'
    )
    # Если каких-то значений нет, заменим NaN на 0
    result['Кол-во входов'] = result['Кол-во входов'].fillna(0).astype(int)
    result['Кол-во выходов'] = result['Кол-во выходов'].fillna(0).astype(int)

    # =========================
    # 13. Оставляем только нужные столбцы
    # =========================
    final = result[[
        'Сотрудник',
        'Дата',
        'Время прихода',
        'Время ухода',
        'Часы отсутствия',
        'Переработка',
        'Кол-во входов',
        'Кол-во выходов'
    ]]

    # Для отладки/просмотра в консоли
    pd.set_option('display.max_rows', None)
    pd.set_option('display.width', None)
    print(final)

    # =========================
    # 14. Сохранение результата с шириной столбцов и границами
    # =========================
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    sheet_name = 'Итог'

    if same_file:
        try:
            with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
                final.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                # Установим ширину столбцов A–H
                ws.column_dimensions['A'].width = 35  # Сотрудник
                ws.column_dimensions['B'].width = 12  # Дата
                ws.column_dimensions['C'].width = 15  # Время прихода
                ws.column_dimensions['D'].width = 15  # Время ухода
                ws.column_dimensions['E'].width = 16  # Часы отсутствия
                ws.column_dimensions['F'].width = 14  # Переработка
                ws.column_dimensions['G'].width = 15  # Кол-во входов
                ws.column_dimensions['H'].width = 15  # Кол-во выходов

                # Проставляем рамки во всех заполненных ячейках
                max_row = ws.max_row
                max_col = ws.max_column
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.border = thin_border

            print(f"Результат добавлен в файл '{file_path}' на лист '{sheet_name}' с оформлением.")
        except Exception as e:
            print(f"Ошибка при добавлении листа в '{file_path}': {e}")
    else:
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                final.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                ws.column_dimensions['A'].width = 35
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 16
                ws.column_dimensions['F'].width = 14
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15

                max_row = ws.max_row
                max_col = ws.max_column
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.border = thin_border

            print(f"Результат сохранён в файл '{output_path}' с оформлением.")
        except Exception as e:
            print(f"Ошибка при сохранении файла '{output_path}': {e}")
